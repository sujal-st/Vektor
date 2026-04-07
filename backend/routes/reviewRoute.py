from fastapi import APIRouter, HTTPException, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from database import db
from bson import ObjectId
from jose import jwt, JWTError
from dotenv import load_dotenv
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import os
import openpyxl
import pandas as pd
import sys
from pathlib import Path

# Add backend directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from sentiment_analyzer import clean_text, get_trained_model

load_dotenv()

router = APIRouter()
security = HTTPBearer(auto_error=False)
JWT_SECRET = os.getenv("JWT_SECRET")
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
REVIEWS_XLSX_PATH = os.path.join(BASE_DIR, "data", "new_reviews.xlsx")


class ReviewSchema(BaseModel):
    product_id: str
    text: str
    rating: int  # 1-5

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        if not credentials:
            return None
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=["HS256"])
        return payload
    except JWTError:
        raise HTTPException(status_code=403, detail="Invalid token")


def ensure_admin(user: dict):
    if not user:
        raise HTTPException(status_code=401, detail="Login required")
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admins only")


def load_or_create_workbook():
    if os.path.exists(REVIEWS_XLSX_PATH):
        wb = openpyxl.load_workbook(REVIEWS_XLSX_PATH)
    else:
        os.makedirs(os.path.dirname(REVIEWS_XLSX_PATH), exist_ok=True)
        wb = openpyxl.Workbook()

    sheet = wb.active
    if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
        sheet.append(["id", "user_id", "rating", "text", "product_id", "created_at"])

    return wb, sheet


# check if user has purchased the product
async def has_purchased(user_id: str, product_id: str) -> bool:
    order = await db.orders.find_one({
        "user_id": user_id,
        "items.id": product_id  # check if product exists in any order
    })
    return order is not None


# post a review
@router.post("/reviews")
async def post_review(review: ReviewSchema, user=Depends(get_current_user)):
    try:
        print("user:", user)
        print("review:", review)

        if not user:
            raise HTTPException(status_code=401, detail="Login required to post a review")

        user_id = user["id"]
        user_name = user["userName"]
        print("user id:", user_id)
        print("user name:", user_name)


        # validate rating
        if review.rating < 1 or review.rating > 5:
            raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")

        # check if product exists
        product = await db.products.find_one({"_id": ObjectId(review.product_id)})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        # check if user already reviewed this product
        existing_review = await db.reviews.find_one({
            "user_id": user_id,
            "product_id": review.product_id
        })
        if existing_review:
            raise HTTPException(status_code=400, detail="You have already reviewed this product")

        # check purchase verification
        purchased = await has_purchased(user_id, review.product_id)

        if not purchased:  # ← add this to block
            raise HTTPException(status_code=403, detail="You can only review products you have purchased")

        new_review = {
            "user_id": user_id,
            "user_name": user_name,
            "product_id": review.product_id,
            "text": review.text,
            "rating": review.rating,
            "purchased_verification": purchased,  # True if user bought the product
            "created_at": datetime.utcnow()
        }

        result = await db.reviews.insert_one(new_review)
        new_review["id"] = str(result.inserted_id)
        await save_review_to_excel(new_review)

        return {"message": "Review posted successfully", "review": new_review}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# get all reviews for a product
@router.get("/reviews/{product_id}")
async def get_reviews(product_id: str):
    try:
        reviews = await db.reviews.find(
            {"product_id": product_id}
        ).sort("created_at", -1).to_list(100)  # newest first

        for r in reviews:
            r["id"] = str(r["_id"])
            del r["_id"]

        return reviews

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# delete a review (only by the user who posted it)
@router.delete("/reviews/{review_id}")
async def delete_review(review_id: str, user=Depends(get_current_user)):
    try:
        if not user:
            raise HTTPException(status_code=401, detail="Login required")

        review = await db.reviews.find_one({"_id": ObjectId(review_id)})

        if not review:
            raise HTTPException(status_code=404, detail="Review not found")

        # only allow the owner or admin to delete
        if review["user_id"] != user["id"] and not user.get("is_admin"):
            raise HTTPException(status_code=403, detail="Not authorized to delete this review")

        await db.reviews.delete_one({"_id": ObjectId(review_id)})
        return {"message": "Review deleted successfully"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/reviews/backfill-excel")
async def backfill_reviews_to_excel(user=Depends(get_current_user)):
    try:
        ensure_admin(user)

        wb, sheet = load_or_create_workbook()
        existing_ids = {
            str(cell[0]).strip()
            for cell in sheet.iter_rows(min_row=2, values_only=True)
            if cell and cell[0] is not None
        }

        reviews = await db.reviews.find({}).to_list(None)
        added_count = 0

        for review in reviews:
            review_id = str(review.get("_id"))
            if review_id in existing_ids:
                continue

            sheet.append([
                review_id,
                review.get("user_id"),
                review.get("rating"),
                review.get("text"),
                review.get("product_id"),
                str(review.get("created_at")),
            ])
            existing_ids.add(review_id)
            added_count += 1

        wb.save(REVIEWS_XLSX_PATH)
        return {
            "message": "Backfill completed",
            "added": added_count,
            "total_reviews": len(reviews),
            "file": REVIEWS_XLSX_PATH,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



async def save_review_to_excel(review: dict):
    wb, sheet = load_or_create_workbook()

    sheet.append([
        str(review.get("id")),
        review.get("user_id"),
        review.get("rating"),
        review.get("text"),
        review.get("product_id"),
        str(review.get("created_at")),
    ])
    
    wb.save(REVIEWS_XLSX_PATH)


@router.get("/reviews/{product_id}/sentiment-analysis")
async def analyze_product_sentiment(product_id: str):
    """
    Analyze sentiment of reviews for a specific product using the trained Naive Bayes model.
    Returns breakdown of positive, negative, and neutral reviews with percentages.
    """
    try:
        # Check if Excel file exists
        if not os.path.exists(REVIEWS_XLSX_PATH):
            raise HTTPException(status_code=404, detail="Reviews file not found. Run backfill first.")

        # Load reviews from Excel with header=0 to use first row as column names
        df = pd.read_excel(REVIEWS_XLSX_PATH, header=0)
        
        # If columns are still unnamed (Unnamed: 0, etc), assign expected column names
        expected_columns = ["id", "user_id", "rating", "text", "product_id", "created_at"]
        if any("Unnamed" in col for col in df.columns):
            # Rename columns by position to match expected order
            df.columns = expected_columns[:len(df.columns)]
        
        # Validate required columns
        if 'product_id' not in df.columns or 'text' not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Excel must contain 'product_id' and 'text' columns. Found: {list(df.columns)}"
            )

        # Filter reviews for specific product
        product_df = df[df['product_id'].astype(str) == str(product_id)].copy()
        
        if product_df.empty:
            return {
                "product_id": product_id,
                "total_reviews": 0,
                "positive": 0,
                "negative": 0,
                "neutral": 0,
                "positive_pct": 0.0,
                "negative_pct": 0.0,
                "neutral_pct": 0.0,
                "message": f"No reviews found for product {product_id}"
            }

        # Load trained model
        model, idf = get_trained_model()
        
        # Classify each review
        sentiments = []
        for review_text in product_df['text'].astype(str):
            cleaned_words = clean_text(review_text)
            sentiment = model.predict(cleaned_words, idf)
            sentiments.append(sentiment)

        # Count sentiments
        sentiment_series = pd.Series(sentiments)
        sentiment_counts = sentiment_series.value_counts()
        total = len(sentiments)

        pos = int(sentiment_counts.get('positive', 0))
        neg = int(sentiment_counts.get('negative', 0))
        neu = int(sentiment_counts.get('neutral', 0))

        return {
            "product_id": product_id,
            "total_reviews": total,
            "positive": pos,
            "negative": neg,
            "neutral": neu,
            "positive_pct": round((pos / total) * 100, 2) if total > 0 else 0.0,
            "negative_pct": round((neg / total) * 100, 2) if total > 0 else 0.0,
            "neutral_pct": round((neu / total) * 100, 2) if total > 0 else 0.0,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sentiment analysis failed: {str(e)}")
